"""수집된 JSON 데이터를 세종도서관 제출 양식 Excel(.xlsx)로 변환"""
import csv
import json
import glob
import os
import re
from collections import OrderedDict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from site_manager import SiteManager
from brm_mapper import BrmMapper
from relevance_filter import is_excluded_topic, is_low_quality, deduplicate_results

DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
OUTPUT_DIR = DATA_DIR
COMPLETED_CSV = os.path.join(os.path.dirname(__file__), "completed sites.csv")

# 수집 대상 최소 날짜 (2026년 1월 1일 이후만 수집)
MIN_DATE = datetime(2026, 1, 1)

# 세종도서관 제출 양식 23컬럼 (참고문서 수집 양식 기준)
HEADERS = [
    "No", "국내/외", "국가", "자료명", "발행처", "저자", "발행연도",
    "BRM코드1", "BRM코드2", "BRM코드1", "BRM코드2",
    "식별번호", "URL", "첨부파일출처", "수록잡지", "권호정보",
    "회의명", "회의일시", "장소", "주최", "공포일자", "발췌", "키워드", "라이선스",
]


def parse_date(date_str):
    """발행일 문자열을 datetime으로 파싱. 실패 시 None 반환."""
    if not date_str:
        return None
    # ISO 형식: "2026-02-19T13:35:35Z" or "2026-02-19"
    for fmt in ("%Y-%m-%dT%H:%M:%SZ", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(date_str.strip(), fmt)
        except ValueError:
            pass
    # 영문 월 형식: "Feb 19, 2026"
    try:
        return datetime.strptime(date_str.strip(), "%b %d, %Y")
    except ValueError:
        pass
    return None


def extract_year(date_str):
    """발행일 문자열에서 4자리 연도 추출"""
    match = re.search(r"(\d{4})", date_str or "")
    if match:
        year = int(match.group(1))
        if 1900 <= year <= 2100:
            return str(year)
    return ""


def load_completed_set():
    """completed sites.csv에서 (발행처, 자료명) 세트를 로드"""
    completed = set()
    if not os.path.exists(COMPLETED_CSV):
        return completed
    with open(COMPLETED_CSV, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            publisher = (row.get("발행처") or "").strip()
            title = (row.get("자료명") or "").strip()
            if publisher and title:
                completed.add((publisher, title))
    return completed


def load_all_documents():
    """data/ 폴더의 모든 JSON 파일에서 문서를 로드 (세종도서관 양식)"""
    sm = SiteManager()
    brm = BrmMapper()

    # 기존 완료 자료 로드 (중복 제거용)
    completed = load_completed_set()
    skipped_dup = 0
    skipped_old = 0
    skipped_no_pdf = 0
    skipped_excluded = 0
    skipped_low_q = 0

    documents = []
    json_files = sorted(glob.glob(os.path.join(DATA_DIR, "*.json")))

    for fpath in json_files:
        fname = os.path.basename(fpath)
        # summary 파일 제외
        if "summary" in fname:
            continue
        with open(fpath, "r", encoding="utf-8") as f:
            data = json.load(f)

        meta = data.get("metadata", {})
        site_code = meta.get("site_code", "")

        # SiteManager에서 사이트 조회
        site = sm.get_by_code(site_code)
        country_code = site.country_code if site else ""
        site_name = meta.get("site_name", "")

        # BRM 코드 매핑
        brm_codes = brm.get_brm_for_site(site) if site else []
        brm1_code1 = brm_codes[0][0] if len(brm_codes) >= 1 else ""
        brm1_code2 = brm_codes[0][1] if len(brm_codes) >= 1 else ""
        brm2_code1 = brm_codes[1][0] if len(brm_codes) >= 2 else ""
        brm2_code2 = brm_codes[1][1] if len(brm_codes) >= 2 else ""

        for doc in data.get("documents", []):
            title = doc.get("title", "").strip()

            # completed sites.csv에 이미 있는 자료는 제외
            if (site_name, title) in completed:
                skipped_dup += 1
                continue

            # 2025년 12월 이전 자료 제외 (날짜 있을 때만 필터)
            pub_date_str = doc.get("published_date", "")
            pub_date = parse_date(pub_date_str)
            if pub_date and pub_date < MIN_DATE:
                skipped_old += 1
                continue

            # 비정책 주제 / 품질 미달 필터 (원본 doc 기준)
            if is_excluded_topic(doc):
                skipped_excluded += 1
                continue
            if is_low_quality(doc):
                skipped_low_q += 1
                continue

            report_id = (doc.get("report_id") or doc.get("report_number") or "").strip()

            row = {
                "국내/외": "1",  # 모든 수집자료는 국외(1)
                "국가": country_code,
                "자료명": title,
                "발행처": site_name,
                "저자": doc.get("authors", ""),
                "발행연도": extract_year(doc.get("published_date", "")),
                "BRM코드1_1": brm1_code1,
                "BRM코드2_1": brm1_code2,
                "BRM코드1_2": brm2_code1,
                "BRM코드2_2": brm2_code2,
                "식별번호": report_id if report_id else "N.A",
                "URL": doc.get("link", ""),
                "첨부파일출처": doc.get("pdf_url", ""),
                "수록잡지": doc.get("journal", ""),
                "권호정보": doc.get("volume_info", ""),
                "회의명": doc.get("conference_name", ""),
                "회의일시": doc.get("conference_date", ""),
                "장소": doc.get("venue", ""),
                "주최": doc.get("organizer", ""),
                "공포일자": doc.get("announced_date", ""),
                "발췌": "",  # 요약/번역 시스템에서 채움
                "키워드": doc.get("keywords", ""),
                "라이선스": doc.get("license", ""),
                # 요약 시트용 추가 필드
                "_site_code": site_code,
                "_acronym": (site.acronym if site else "") or meta.get("acronym", ""),
            }
            # PDF URL 없는 문서 제외
            if not doc.get("pdf_url", "").strip():
                skipped_no_pdf += 1
                continue

            documents.append(row)

    if skipped_dup:
        print(f"  (completed sites.csv 중복 {skipped_dup}건 제외)")
    if skipped_old:
        print(f"  (2026-01 이전 자료 {skipped_old}건 제외)")
    if skipped_excluded:
        print(f"  (비정책 주제 {skipped_excluded}건 제외)")
    if skipped_low_q:
        print(f"  (품질 미달 {skipped_low_q}건 제외)")
    if skipped_no_pdf:
        print(f"  (PDF URL 없는 자료 {skipped_no_pdf}건 제외)")

    return documents


def create_xlsx(documents, output_path):
    """문서 리스트를 세종도서관 양식 Excel 파일로 저장"""
    wb = Workbook()

    # ── 전체 목록 시트 ──
    ws = wb.active
    ws.title = "수집자료 전체"

    # 스타일
    header_font = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="맑은 고딕", size=10)
    cell_align = Alignment(vertical="top", wrap_text=True)
    link_font = Font(name="맑은 고딕", size=10, color="0563C1", underline="single")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # 헤더 쓰기
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 컬럼 인덱스 (1-based, 23컬럼)
    COL_NO = 1
    COL_DOMESTIC = 2
    COL_COUNTRY = 3
    COL_TITLE = 4
    COL_PUBLISHER = 5
    COL_AUTHOR = 6
    COL_YEAR = 7
    COL_BRM1_1 = 8
    COL_BRM2_1 = 9
    COL_BRM1_2 = 10
    COL_BRM2_2 = 11
    COL_ID = 12
    COL_URL = 13
    COL_PDF = 14
    COL_JOURNAL = 15
    COL_VOLUME = 16
    COL_CONFERENCE = 17
    COL_CONF_DATE = 18
    COL_VENUE = 19
    COL_ORGANIZER = 20
    COL_ANNOUNCE_DATE = 21
    COL_EXCERPT = 22
    COL_KEYWORD = 23
    COL_LICENSE = 24

    # 데이터 쓰기
    for row_idx, doc in enumerate(documents, 2):
        no = row_idx - 1
        ws.cell(row=row_idx, column=COL_NO, value=no).font = cell_font
        ws.cell(row=row_idx, column=COL_DOMESTIC, value=doc["국내/외"]).font = cell_font
        ws.cell(row=row_idx, column=COL_COUNTRY, value=doc["국가"]).font = cell_font

        # 자료명 (하이퍼링크)
        title_cell = ws.cell(row=row_idx, column=COL_TITLE, value=doc["자료명"])
        url = doc["URL"]
        if url:
            title_cell.hyperlink = url
            title_cell.font = link_font
        else:
            title_cell.font = cell_font

        ws.cell(row=row_idx, column=COL_PUBLISHER, value=doc["발행처"]).font = cell_font
        ws.cell(row=row_idx, column=COL_AUTHOR, value=doc["저자"]).font = cell_font
        ws.cell(row=row_idx, column=COL_YEAR, value=doc["발행연도"]).font = cell_font
        ws.cell(row=row_idx, column=COL_BRM1_1, value=doc["BRM코드1_1"]).font = cell_font
        ws.cell(row=row_idx, column=COL_BRM2_1, value=doc["BRM코드2_1"]).font = cell_font
        ws.cell(row=row_idx, column=COL_BRM1_2, value=doc["BRM코드1_2"]).font = cell_font
        ws.cell(row=row_idx, column=COL_BRM2_2, value=doc["BRM코드2_2"]).font = cell_font
        ws.cell(row=row_idx, column=COL_ID, value=doc["식별번호"]).font = cell_font

        # URL (하이퍼링크)
        url_cell = ws.cell(row=row_idx, column=COL_URL, value=url)
        if url:
            url_cell.hyperlink = url
            url_cell.font = link_font
        else:
            url_cell.font = cell_font

        # 첨부파일출처 (하이퍼링크)
        pdf_url = doc["첨부파일출처"]
        pdf_cell = ws.cell(row=row_idx, column=COL_PDF, value=pdf_url)
        if pdf_url:
            pdf_cell.hyperlink = pdf_url
            pdf_cell.font = link_font
        else:
            pdf_cell.font = cell_font

        ws.cell(row=row_idx, column=COL_JOURNAL, value=doc["수록잡지"]).font = cell_font
        ws.cell(row=row_idx, column=COL_VOLUME, value=doc["권호정보"]).font = cell_font
        ws.cell(row=row_idx, column=COL_CONFERENCE, value=doc.get("회의명", "")).font = cell_font
        ws.cell(row=row_idx, column=COL_CONF_DATE, value=doc.get("회의일시", "")).font = cell_font
        ws.cell(row=row_idx, column=COL_VENUE, value=doc.get("장소", "")).font = cell_font
        ws.cell(row=row_idx, column=COL_ORGANIZER, value=doc.get("주최", "")).font = cell_font
        ws.cell(row=row_idx, column=COL_ANNOUNCE_DATE, value=doc.get("공포일자", "")).font = cell_font
        ws.cell(row=row_idx, column=COL_EXCERPT, value=doc.get("발췌", "")).font = cell_font
        ws.cell(row=row_idx, column=COL_KEYWORD, value=doc["키워드"]).font = cell_font
        ws.cell(row=row_idx, column=COL_LICENSE, value=doc.get("라이선스", "")).font = cell_font

        for col_idx in range(1, len(HEADERS) + 1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.alignment = cell_align
            c.border = thin_border

    # 열 너비 설정
    col_widths = {
        COL_NO: 6,
        COL_DOMESTIC: 8,
        COL_COUNTRY: 6,
        COL_TITLE: 60,
        COL_PUBLISHER: 35,
        COL_AUTHOR: 25,
        COL_YEAR: 10,
        COL_BRM1_1: 10,
        COL_BRM2_1: 10,
        COL_BRM1_2: 10,
        COL_BRM2_2: 10,
        COL_ID: 18,
        COL_URL: 50,
        COL_PDF: 50,
        COL_JOURNAL: 15,
        COL_VOLUME: 12,
        COL_CONFERENCE: 25,
        COL_CONF_DATE: 15,
        COL_VENUE: 20,
        COL_ORGANIZER: 20,
        COL_ANNOUNCE_DATE: 15,
        COL_EXCERPT: 60,
        COL_KEYWORD: 15,
        COL_LICENSE: 15,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # 필터 설정
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(documents) + 1}"

    # 헤더 고정
    ws.freeze_panes = "A2"

    # ── 요약 시트 ──
    ws_sum = wb.create_sheet("수집 요약")

    # 국가별/기관별 집계
    site_counts = OrderedDict()
    for doc in documents:
        key = (doc["국가"], doc["_site_code"], doc["발행처"], doc["_acronym"],
               doc["BRM코드1_1"], doc["BRM코드2_1"])
        site_counts[key] = site_counts.get(key, 0) + 1

    sum_headers = ["No", "국가", "기관코드", "발행처", "기관약어", "BRM코드1", "BRM코드2", "수집건수"]
    for col_idx, header in enumerate(sum_headers, 1):
        cell = ws_sum.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, ((country, code, name, acronym, brm1, brm2), count) in enumerate(site_counts.items(), 2):
        ws_sum.cell(row=row_idx, column=1, value=row_idx - 1).font = cell_font
        ws_sum.cell(row=row_idx, column=2, value=country).font = cell_font
        ws_sum.cell(row=row_idx, column=3, value=code).font = cell_font
        ws_sum.cell(row=row_idx, column=4, value=name).font = cell_font
        ws_sum.cell(row=row_idx, column=5, value=acronym).font = cell_font
        ws_sum.cell(row=row_idx, column=6, value=brm1).font = cell_font
        ws_sum.cell(row=row_idx, column=7, value=brm2).font = cell_font
        ws_sum.cell(row=row_idx, column=8, value=count).font = cell_font
        for col_idx in range(1, len(sum_headers) + 1):
            ws_sum.cell(row=row_idx, column=col_idx).border = thin_border
            ws_sum.cell(row=row_idx, column=col_idx).alignment = cell_align

    # 합계 행
    total_row = len(site_counts) + 2
    ws_sum.cell(row=total_row, column=7, value="합계").font = Font(name="맑은 고딕", bold=True, size=11)
    ws_sum.cell(row=total_row, column=8, value=len(documents)).font = Font(name="맑은 고딕", bold=True, size=11)

    sum_col_widths = {1: 6, 2: 6, 3: 10, 4: 40, 5: 10, 6: 10, 7: 10, 8: 10}
    for col, width in sum_col_widths.items():
        ws_sum.column_dimensions[get_column_letter(col)].width = width

    ws_sum.auto_filter.ref = f"A1:{get_column_letter(len(sum_headers))}{len(site_counts) + 1}"
    ws_sum.freeze_panes = "A2"

    # 저장
    wb.save(output_path)
    return output_path


def main():
    print("JSON 데이터 로드 중...")
    documents = load_all_documents()
    print(f"총 {len(documents)}건 로드 완료")

    today = datetime.now().strftime("%Y%m%d")
    output_path = os.path.join(OUTPUT_DIR, f"세종도서관_해외자료수집_{today}.xlsx")

    print("Excel 파일 생성 중...")
    result = create_xlsx(documents, output_path)
    print(f"완료: {result}")
    print(f"  - 시트1: 수집자료 전체 ({len(documents)}건, 24컬럼 세종도서관 양식+라이선스)")
    print(f"  - 시트2: 수집 요약 (기관별 집계)")


if __name__ == "__main__":
    main()
