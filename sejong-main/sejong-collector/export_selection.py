"""
selection_330.json에서 최종 300건을 선별하여 세종도서관 양식 Excel로 내보내기.
PDF URL이 없는 경우 랜딩페이지 URL을 첨부파일출처로 사용.
"""
import json
import os
import re
import sys
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from site_manager import SiteManager
from brm_mapper import BrmMapper

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace', line_buffering=True)

DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
SELECTION_FILE = os.path.join(DATA_DIR, "selection_400.json")

HEADERS = [
    "No", "국내/외", "국가", "자료명", "발행처", "저자", "발행연도",
    "BRM코드1", "BRM코드2", "BRM코드1", "BRM코드2",
    "식별번호", "유형", "URL", "첨부파일출처", "수록잡지", "권호정보",
    "회의명", "회의일시", "장소", "주최", "공포일자", "발췌", "키워드", "라이선스",
]


def extract_year(date_str):
    match = re.search(r"(\d{4})", date_str or "")
    if match:
        year = int(match.group(1))
        if 1900 <= year <= 2100:
            return str(year)
    return ""


def main():
    sm = SiteManager()
    brm = BrmMapper()

    with open(SELECTION_FILE, 'r', encoding='utf-8') as f:
        all_docs = json.load(f)

    print(f"Loaded {len(all_docs)} docs from selection")

    # Sort by relevance score desc, take only non-excluded docs
    all_docs.sort(key=lambda x: -x.get('_relevance_score', 0))
    selected = [d for d in all_docs if d.get('_relevance_score', 0) > -1]
    print(f"Selected {len(selected)} verified docs")

    # Stats
    has_pdf = sum(1 for d in selected if d.get('pdf_url'))
    print(f"With PDF: {has_pdf}, Without PDF: {len(selected) - has_pdf}")

    # Build rows
    rows = []
    for doc in selected:
        site_code = doc.get('site_code', '')
        site = sm.get_by_code(site_code) if site_code else None

        country_code = doc.get('_country', '')
        # Map to Sejong Library country codes (non-ISO)
        COUNTRY_CODE_MAP = {
            'se': 'SW',   # 스웨덴
            'sg': 'SI',   # 싱가포르
            'ge': 'DE',   # 독일
            'ch': 'CH',   # 스위스 (not in ref, keep as-is)
            'jo': 'JO',   # 요르단 (not in ref, keep as-is)
            'nz': 'NZ',   # 뉴질랜드 (not in ref, keep as-is)
            'sa': 'SA',   # 사우디 (not in ref, keep as-is)
        }
        country_code = COUNTRY_CODE_MAP.get(country_code.lower(), country_code)
        site_name = doc.get('_site_name', '') or doc.get('site_name', '')

        # BRM codes
        brm_codes = brm.get_brm_for_site(site) if site else []
        # Fallback BRM mapping from matched keywords
        if not brm_codes:
            KW_TO_BRM = {
                'AI_디지털_반도체': [('0002', '0030')],      # 과학기술 > 과학기술연구
                '기후_에너지': [('0017', '0221')],            # 환경 > 환경일반
                '경제_통상': [('0011', '0159')],              # 산업통상 > 통상
                '안보_외교': [('0015', '0199')],              # 통일외교 > 외교
                '인구_복지': [('0010', '0154')],              # 사회복지 > 사회복지일반
                '보건': [('0009', '0125')],                   # 보건 > 보건의료
                '주거_도시': [('0006', '0067')],              # 지역개발 > 지역및도시
                '교육': [('0003', '0039')],                   # 교육 > 교육일반
            }
            matched = doc.get('_matched_keywords', [])
            for m in matched:
                if m in KW_TO_BRM:
                    brm_codes = KW_TO_BRM[m]
                    break
        brm1_code1 = brm_codes[0][0] if len(brm_codes) >= 1 else ""
        brm1_code2 = brm_codes[0][1] if len(brm_codes) >= 1 else ""
        brm2_code1 = brm_codes[1][0] if len(brm_codes) >= 2 else ""
        brm2_code2 = brm_codes[1][1] if len(brm_codes) >= 2 else ""

        title = doc.get('title', '').strip()
        url = doc.get('link', '')
        pdf_url = doc.get('pdf_url', '') or url  # Fallback to landing page
        report_id = (doc.get('report_id') or doc.get('report_number') or '').strip()

        # 식별번호: ISBN 우선, ISSN 차선, report_id 폴백
        isbn = (doc.get('isbn') or '').strip()
        issn = (doc.get('issn') or '').strip()
        if isbn:
            identifier = isbn
        elif issn:
            identifier = issn
        elif report_id:
            identifier = report_id
        else:
            identifier = ""

        # 유형
        doc_type = doc.get('doc_type', '') or '발간자료'

        row = {
            "국내/외": "1",
            "국가": country_code.upper(),
            "자료명": title,
            "발행처": site_name,
            "저자": doc.get('authors', ''),
            "발행연도": extract_year(doc.get('published_date', '')),
            "BRM코드1_1": brm1_code1,
            "BRM코드2_1": brm1_code2,
            "BRM코드1_2": brm2_code1,
            "BRM코드2_2": brm2_code2,
            "식별번호": identifier,
            "유형": doc_type,
            "URL": url,
            "첨부파일출처": pdf_url,
            "수록잡지": doc.get('journal', ''),
            "권호정보": doc.get('volume_info', ''),
            "회의명": "",
            "회의일시": "",
            "장소": "",
            "주최": "",
            "공포일자": "",
            "발췌": doc.get('description', ''),
            "키워드": doc.get('keywords', ''),
            "라이선스": doc.get('license', ''),
        }
        rows.append(row)

    # Create Excel
    today = datetime.now().strftime('%Y%m%d')
    output_path = os.path.join(DATA_DIR, f"세종도서관_해외자료수집_{today}_v2.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "수집자료 전체"

    # Styles
    header_font = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="맑은 고딕", size=10)
    cell_align = Alignment(vertical="top", wrap_text=True)
    link_font = Font(name="맑은 고딕", size=10, color="0563C1", underline="single")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # Headers
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    for row_idx, doc in enumerate(rows, 2):
        no = row_idx - 1
        ws.cell(row=row_idx, column=1, value=no).font = cell_font
        ws.cell(row=row_idx, column=2, value=doc["국내/외"]).font = cell_font
        ws.cell(row=row_idx, column=3, value=doc["국가"]).font = cell_font

        title_cell = ws.cell(row=row_idx, column=4, value=doc["자료명"])
        url = doc["URL"]
        if url:
            try:
                title_cell.hyperlink = url
                title_cell.font = link_font
            except Exception:
                title_cell.font = cell_font
        else:
            title_cell.font = cell_font

        ws.cell(row=row_idx, column=5, value=doc["발행처"]).font = cell_font
        ws.cell(row=row_idx, column=6, value=doc["저자"]).font = cell_font
        ws.cell(row=row_idx, column=7, value=doc["발행연도"]).font = cell_font
        ws.cell(row=row_idx, column=8, value=doc["BRM코드1_1"]).font = cell_font
        ws.cell(row=row_idx, column=9, value=doc["BRM코드2_1"]).font = cell_font
        ws.cell(row=row_idx, column=10, value=doc["BRM코드1_2"]).font = cell_font
        ws.cell(row=row_idx, column=11, value=doc["BRM코드2_2"]).font = cell_font
        ws.cell(row=row_idx, column=12, value=doc["식별번호"]).font = cell_font
        ws.cell(row=row_idx, column=13, value=doc["유형"]).font = cell_font

        url_cell = ws.cell(row=row_idx, column=14, value=doc["URL"])
        if doc["URL"]:
            try:
                url_cell.hyperlink = doc["URL"]
                url_cell.font = link_font
            except Exception:
                url_cell.font = cell_font

        pdf_cell = ws.cell(row=row_idx, column=15, value=doc["첨부파일출처"])
        if doc["첨부파일출처"]:
            try:
                pdf_cell.hyperlink = doc["첨부파일출처"]
                pdf_cell.font = link_font
            except Exception:
                pdf_cell.font = cell_font

        ws.cell(row=row_idx, column=16, value=doc["수록잡지"]).font = cell_font
        ws.cell(row=row_idx, column=17, value=doc["권호정보"]).font = cell_font
        ws.cell(row=row_idx, column=18, value=doc["회의명"]).font = cell_font
        ws.cell(row=row_idx, column=19, value=doc["회의일시"]).font = cell_font
        ws.cell(row=row_idx, column=20, value=doc["장소"]).font = cell_font
        ws.cell(row=row_idx, column=21, value=doc["주최"]).font = cell_font
        ws.cell(row=row_idx, column=22, value=doc["공포일자"]).font = cell_font
        ws.cell(row=row_idx, column=23, value=doc["발췌"]).font = cell_font
        ws.cell(row=row_idx, column=24, value=doc["키워드"]).font = cell_font
        ws.cell(row=row_idx, column=25, value=doc["라이선스"]).font = cell_font

        # Apply borders
        for col in range(1, 26):
            ws.cell(row=row_idx, column=col).border = thin_border
            ws.cell(row=row_idx, column=col).alignment = cell_align

    # Column widths
    col_widths = {
        1: 5, 2: 6, 3: 6, 4: 60, 5: 30, 6: 20, 7: 8,
        8: 8, 9: 8, 10: 8, 11: 8, 12: 18, 13: 10, 14: 40, 15: 40,
        16: 25, 17: 15, 18: 15, 19: 12, 20: 12, 21: 12,
        22: 12, 23: 30, 24: 30, 25: 10,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[chr(64 + col) if col <= 26 else ''].width = width

    # Freeze panes
    ws.freeze_panes = 'A2'

    # Save
    wb.save(output_path)
    print(f"\nExcel saved: {output_path}")
    print(f"Total rows: {len(rows)}")

    # Country summary
    from collections import defaultdict
    country_counts = defaultdict(int)
    for r in rows:
        country_counts[r["국가"]] += 1
    print(f"\nCountry distribution:")
    for k, v in sorted(country_counts.items(), key=lambda x: -x[1]):
        print(f"  {k}: {v}")


if __name__ == '__main__':
    main()
