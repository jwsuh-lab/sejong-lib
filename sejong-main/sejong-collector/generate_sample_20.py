"""PDF URL이 있는 문서 20건 샘플 추출 → 세종도서관 양식 Excel"""
import json, glob, os, sys, csv, re
from datetime import datetime
from collections import defaultdict, OrderedDict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from site_manager import SiteManager
from brm_mapper import BrmMapper

DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
COMPLETED_CSV = os.path.join(os.path.dirname(__file__), "completed sites.csv")
MIN_DATE = datetime(2025, 12, 1)

HEADERS = [
    "No", "국내/외", "국가", "자료명", "발행처", "저자", "발행연도",
    "BRM코드1", "BRM코드2", "BRM코드1", "BRM코드2",
    "식별번호", "URL", "첨부파일출처", "수록잡지", "권호정보", "키워드",
]


def parse_date(date_str):
    if not date_str:
        return None
    for fmt in ("%Y-%m-%dT%H:%M:%SZ", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(date_str.strip(), fmt)
        except ValueError:
            pass
    try:
        from dateutil import parser as dp
        return dp.parse(date_str).replace(tzinfo=None)
    except Exception:
        pass
    return None


def load_completed():
    completed = set()
    if not os.path.exists(COMPLETED_CSV):
        return completed
    with open(COMPLETED_CSV, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            publisher = (row.get("발행처") or "").strip()
            title = (row.get("자료명") or "").strip()
            if publisher and title:
                completed.add((publisher, title))
    return completed


def collect_pdf_docs():
    """모든 JSON 파일에서 PDF URL 있는 문서 수집 (필터 적용)"""
    sm = SiteManager()
    brm_mapper = BrmMapper()
    completed = load_completed()

    all_pdf_docs = []
    seen_titles = set()
    json_files = sorted(glob.glob(os.path.join(DATA_DIR, "*.json")))

    for fpath in json_files:
        fname = os.path.basename(fpath)
        if "summary" in fname:
            continue

        with open(fpath, "r", encoding="utf-8") as f:
            data = json.load(f)

        meta = data.get("metadata", {})
        site_code = meta.get("site_code", "")
        site_name = meta.get("site_name", "")

        site = sm.get_by_code(site_code)
        country_code = site.country_code if site else ""
        brm_codes = brm_mapper.get_brm_for_site(site) if site else []
        brm1_code1 = brm_codes[0][0] if len(brm_codes) >= 1 else ""
        brm1_code2 = brm_codes[0][1] if len(brm_codes) >= 1 else ""
        brm2_code1 = brm_codes[1][0] if len(brm_codes) >= 2 else ""
        brm2_code2 = brm_codes[1][1] if len(brm_codes) >= 2 else ""

        for doc in data.get("documents", []):
            pdf_url = doc.get("pdf_url", "")
            title = doc.get("title", "").strip()

            if not pdf_url or not title:
                continue
            if title in seen_titles:
                continue
            if (site_name, title) in completed:
                continue

            pub_date = parse_date(doc.get("published_date", ""))
            if pub_date and pub_date < MIN_DATE:
                continue

            seen_titles.add(title)

            report_id = (doc.get("report_id") or doc.get("report_number") or "").strip()
            year_match = re.search(r"(\d{4})", doc.get("published_date", "") or "")
            year = ""
            if year_match:
                yr = int(year_match.group(1))
                if 1900 <= yr <= 2100:
                    year = str(yr)

            row = {
                "국내/외": "1",
                "국가": country_code,
                "자료명": title,
                "발행처": site_name,
                "저자": doc.get("authors", ""),
                "발행연도": year,
                "BRM코드1_1": brm1_code1,
                "BRM코드2_1": brm1_code2,
                "BRM코드1_2": brm2_code1,
                "BRM코드2_2": brm2_code2,
                "식별번호": report_id if report_id else "N.A",
                "URL": doc.get("link", ""),
                "첨부파일출처": pdf_url,
                "수록잡지": doc.get("journal", ""),
                "권호정보": doc.get("volume_info", ""),
                "키워드": doc.get("keywords", ""),
                "_site_code": site_code,
                "_acronym": (site.acronym if site else "") or meta.get("acronym", ""),
            }
            all_pdf_docs.append({"site_code": site_code, "row": row})

    return all_pdf_docs


def select_20_diverse(all_pdf_docs):
    """사이트별 라운드로빈으로 20건 선택"""
    by_site = defaultdict(list)
    for d in all_pdf_docs:
        by_site[d["site_code"]].append(d["row"])

    site_keys = list(by_site.keys())
    selected = []
    round_num = 0

    while len(selected) < 20:
        added = False
        for code in site_keys:
            docs = by_site[code]
            if round_num < len(docs):
                selected.append(docs[round_num])
                added = True
                if len(selected) >= 20:
                    break
        if not added:
            break
        round_num += 1

    return selected


def create_xlsx(documents, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "수집자료 전체"

    # Styles
    header_font = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="맑은 고딕", size=10)
    cell_align = Alignment(vertical="top", wrap_text=True)
    link_font = Font(name="맑은 고딕", size=10, color="0563C1", underline="single")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # Header
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    COL = {
        "NO": 1, "DOMESTIC": 2, "COUNTRY": 3, "TITLE": 4, "PUBLISHER": 5,
        "AUTHOR": 6, "YEAR": 7, "BRM1_1": 8, "BRM2_1": 9, "BRM1_2": 10,
        "BRM2_2": 11, "ID": 12, "URL": 13, "PDF": 14, "JOURNAL": 15,
        "VOLUME": 16, "KEYWORD": 17,
    }

    for row_idx, doc in enumerate(documents, 2):
        no = row_idx - 1
        ws.cell(row=row_idx, column=COL["NO"], value=no).font = cell_font
        ws.cell(row=row_idx, column=COL["DOMESTIC"], value=doc["국내/외"]).font = cell_font
        ws.cell(row=row_idx, column=COL["COUNTRY"], value=doc["국가"]).font = cell_font

        title_cell = ws.cell(row=row_idx, column=COL["TITLE"], value=doc["자료명"])
        url = doc["URL"]
        if url:
            title_cell.hyperlink = url
            title_cell.font = link_font
        else:
            title_cell.font = cell_font

        ws.cell(row=row_idx, column=COL["PUBLISHER"], value=doc["발행처"]).font = cell_font
        ws.cell(row=row_idx, column=COL["AUTHOR"], value=doc["저자"]).font = cell_font
        ws.cell(row=row_idx, column=COL["YEAR"], value=doc["발행연도"]).font = cell_font
        ws.cell(row=row_idx, column=COL["BRM1_1"], value=doc["BRM코드1_1"]).font = cell_font
        ws.cell(row=row_idx, column=COL["BRM2_1"], value=doc["BRM코드2_1"]).font = cell_font
        ws.cell(row=row_idx, column=COL["BRM1_2"], value=doc["BRM코드1_2"]).font = cell_font
        ws.cell(row=row_idx, column=COL["BRM2_2"], value=doc["BRM코드2_2"]).font = cell_font
        ws.cell(row=row_idx, column=COL["ID"], value=doc["식별번호"]).font = cell_font

        url_cell = ws.cell(row=row_idx, column=COL["URL"], value=url)
        if url:
            url_cell.hyperlink = url
            url_cell.font = link_font
        else:
            url_cell.font = cell_font

        pdf_url = doc["첨부파일출처"]
        pdf_cell = ws.cell(row=row_idx, column=COL["PDF"], value=pdf_url)
        if pdf_url:
            pdf_cell.hyperlink = pdf_url
            pdf_cell.font = link_font
        else:
            pdf_cell.font = cell_font

        ws.cell(row=row_idx, column=COL["JOURNAL"], value=doc["수록잡지"]).font = cell_font
        ws.cell(row=row_idx, column=COL["VOLUME"], value=doc["권호정보"]).font = cell_font
        ws.cell(row=row_idx, column=COL["KEYWORD"], value=doc["키워드"]).font = cell_font

        for ci in range(1, len(HEADERS) + 1):
            c = ws.cell(row=row_idx, column=ci)
            c.alignment = cell_align
            c.border = thin_border

    # Column widths
    col_widths = {
        1: 6, 2: 8, 3: 6, 4: 60, 5: 35, 6: 25, 7: 10,
        8: 10, 9: 10, 10: 10, 11: 10, 12: 18, 13: 50, 14: 50,
        15: 25, 16: 12, 17: 35,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(documents) + 1}"
    ws.freeze_panes = "A2"

    # Summary sheet
    ws_sum = wb.create_sheet("수집 요약")
    site_counts = OrderedDict()
    for doc in documents:
        key = (doc["국가"], doc["_site_code"], doc["발행처"], doc["_acronym"],
               doc["BRM코드1_1"], doc["BRM코드2_1"])
        site_counts[key] = site_counts.get(key, 0) + 1

    sum_headers = ["No", "국가", "기관코드", "발행처", "기관약어", "BRM코드1", "BRM코드2", "수집건수"]
    for col_idx, header in enumerate(sum_headers, 1):
        cell = ws_sum.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, ((country, code, name, acronym, brm1, brm2), count) in enumerate(site_counts.items(), 2):
        ws_sum.cell(row=row_idx, column=1, value=row_idx - 1).font = cell_font
        ws_sum.cell(row=row_idx, column=2, value=country).font = cell_font
        ws_sum.cell(row=row_idx, column=3, value=code).font = cell_font
        ws_sum.cell(row=row_idx, column=4, value=name).font = cell_font
        ws_sum.cell(row=row_idx, column=5, value=acronym).font = cell_font
        ws_sum.cell(row=row_idx, column=6, value=brm1).font = cell_font
        ws_sum.cell(row=row_idx, column=7, value=brm2).font = cell_font
        ws_sum.cell(row=row_idx, column=8, value=count).font = cell_font
        for ci in range(1, len(sum_headers) + 1):
            ws_sum.cell(row=row_idx, column=ci).border = thin_border
            ws_sum.cell(row=row_idx, column=ci).alignment = cell_align

    total_row = len(site_counts) + 2
    ws_sum.cell(row=total_row, column=7, value="합계").font = Font(name="맑은 고딕", bold=True, size=11)
    ws_sum.cell(row=total_row, column=8, value=len(documents)).font = Font(name="맑은 고딕", bold=True, size=11)

    sum_col_widths = {1: 6, 2: 6, 3: 10, 4: 40, 5: 10, 6: 10, 7: 10, 8: 10}
    for col, width in sum_col_widths.items():
        ws_sum.column_dimensions[get_column_letter(col)].width = width

    ws_sum.auto_filter.ref = f"A1:{get_column_letter(len(sum_headers))}{len(site_counts) + 1}"
    ws_sum.freeze_panes = "A2"

    wb.save(output_path)
    return output_path


def main():
    print("PDF URL 있는 문서 수집 중...")
    all_pdf_docs = collect_pdf_docs()
    print(f"  PDF URL 보유 문서: {len(all_pdf_docs)}건 (12개 사이트)")

    print("\n20건 다양한 사이트 라운드로빈 선택...")
    selected = select_20_diverse(all_pdf_docs)
    print(f"  선택: {len(selected)}건")

    # 선택된 문서 목록
    for i, doc in enumerate(selected, 1):
        print(f"  {i:2d}. [{doc['_site_code']}] {doc['발행처'][:25]:25s} | {doc['자료명'][:40]}")

    # Excel 생성
    output_path = os.path.join(DATA_DIR, "세종도서관_메타데이터샘플_20건_20260224.xlsx")
    print(f"\nExcel 생성 중...")
    create_xlsx(selected, output_path)
    print(f"완료: {output_path}")

    # 메타데이터 채워진 현황
    journal_filled = sum(1 for d in selected if d["수록잡지"])
    keyword_filled = sum(1 for d in selected if d["키워드"])
    volume_filled = sum(1 for d in selected if d["권호정보"])
    pdf_filled = sum(1 for d in selected if d["첨부파일출처"])

    print(f"\n=== 메타데이터 채워진 현황 ===")
    print(f"  수록잡지:    {journal_filled}/20")
    print(f"  키워드:      {keyword_filled}/20")
    print(f"  권호정보:    {volume_filled}/20")
    print(f"  첨부파일출처: {pdf_filled}/20")


if __name__ == "__main__":
    main()
